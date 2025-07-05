import streamlit as st
import pandas as pd
import zipfile
import tempfile
import os
from openpyxl import load_workbook

def read_file_a(file_a_path):
    df = pd.read_excel(file_a_path, index_col=0)
    wage_cols = df.columns
    positive_or_zero = {}
    negative = {}

    for budget_unit, row in df.iterrows():
        for wage_type in wage_cols:
            value = row[wage_type]
            if "绩效工资" in wage_type:
                wage_type = wage_type.replace("绩效工资", "基础性绩效")
            if "行政医疗" in wage_type:
                wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
            elif "事业医疗" in wage_type:
                wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
            elif "医疗保险" in wage_type:
                wage_type = wage_type.replace("医疗保险", "基本医疗")
            key = (str(budget_unit).strip(), str(wage_type).strip())
            if value < 0:
                negative[key] = value
            else:
                positive_or_zero[key] = value
    return positive_or_zero, negative

def update_template(template_path, value_dict, unit_col, type_col, value_col, output_name):
    wb = load_workbook(template_path)
    sheet = wb.active
    match_count = 0

    for row_idx in range(2, sheet.max_row + 1):
        unit = str(sheet.cell(row=row_idx, column=unit_col).value or "").strip()
        wage_type = str(sheet.cell(row=row_idx, column=type_col).value or "").strip()
        unit_cleaned = unit.replace("-", "").replace(" ", "")
        for (k_unit, k_type), value in value_dict.items():
            k_unit_cleaned = k_unit.replace("-", "").replace(" ", "")
            if (k_unit_cleaned in unit_cleaned or unit_cleaned in k_unit_cleaned) and k_type in wage_type:
                sheet.cell(row=row_idx, column=value_col).value = value
                match_count += 1
                break

    output_path = os.path.join(os.path.dirname(template_path), output_name)
    wb.save(output_path)
    return output_path, match_count

# Streamlit 网页界面
st.title("工资数据自动填充工具")
uploaded_zip = st.file_uploader("请上传包含 文件A、模板A、模板B 的压缩包 (.zip)", type="zip")

if uploaded_zip:
    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, "uploaded.zip")
        with open(zip_path, "wb") as f:
            f.write(uploaded_zip.read())

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(tmpdir)

        # 尝试查找文件路径
        files = os.listdir(tmpdir)
        file_a_path = next((os.path.join(tmpdir, f) for f in files if "文件A" in f), None)
        template_a_path = next((os.path.join(tmpdir, f) for f in files if "模板A" in f), None)
        template_b_path = next((os.path.join(tmpdir, f) for f in files if "模板B" in f), None)

        if file_a_path and template_a_path and template_b_path:
            pos_zero_dict, neg_dict = read_file_a(file_a_path)
            output_a, count_a = update_template(template_a_path, pos_zero_dict, 1, 2, 10, "updated_模板A.xlsx")
            output_b, count_b = update_template(template_b_path, neg_dict, 2, 3, 7, "updated_模板B.xlsx")

            st.success(f"模板A已更新，共匹配 {count_a} 项")
            st.download_button("下载 updated_模板A.xlsx", data=open(output_a, "rb").read(), file_name="updated_模板A.xlsx")

            st.success(f"模板B已更新，共匹配 {count_b} 项")
            st.download_button("下载 updated_模板B.xlsx", data=open(output_b, "rb").read(), file_name="updated_模板B.xlsx")
        else:
            st.error("压缩包内未找到所需的3个文件（文件A、模板A、模板B）")
